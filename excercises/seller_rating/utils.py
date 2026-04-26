import openpyxl as xl

# reward_dict = {
#     "Gold": { "starting": 500001, "ending": 10000000 },
#     "Silver": {"starting": 250001, "ending": 500000 },
#     "Bronze": { "starting": 100001, "ending": 250000 },
#     "No Rewards": {"starting": 0, "ending": 100000 }
# }

reward_dict_array = [
    {"id": 1, "name": "Gold", "starting": 500001, "ending": 10000000 },
    {"id": 2, "name": "Silver", "starting": 250001, "ending": 500000 },
    {"id": 3, "name": "Bronze", "starting": 100001, "ending": 250000 },
    {"id": 4, "name": "No Rewards", "starting": 0, "ending": 100000 }
]

def get_reward_by_profit(profit):
    #filtered_dict = {k: v for k, v in reward_dict.items() if (profit >= v['starting'] and profit <= v['ending'])}
    #reward = {list(filtered_dict)[0]}
    matched_reward = ''
    for reward in reward_dict_array:
        if profit >= reward['starting'] and profit <= reward['ending']:
            matched_reward = reward['name']
            break
    return matched_reward

def write_data_to_excel(sellers_array):
    # wb = xl.load_workbook('sample_data_large.xlsx')
    # sheet = wb['Sheet1']
    # cell = sheet['a1'] # or cell = sheet.cell(1,1)
    # print(cell,value)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sellers"

    # 4. Write headers to the first row
    headers = ["Id", "Name", "Total Profit", "Reward"]

    # 5. Write the class data
    cell_row = 1
    ws.append(headers)
    cell_row += 1
    for seller in sellers_array:
        new_call = ws.cell(row=cell_row, column=1)
        new_call.value = seller.id
        new_call = ws.cell(row=cell_row, column=2)
        new_call.value = seller.name
        new_call = ws.cell(row=cell_row, column=3)
        new_call.value = seller.total_profit
        new_call = ws.cell(row=cell_row, column=4)
        new_call.value = str(seller.reward)

        cell_row += 1


    # 6. Save the workbook
    file_path = "sellers.xlsx"
    wb.save(file_path)

    print(f"Data successfully written to {file_path}")
